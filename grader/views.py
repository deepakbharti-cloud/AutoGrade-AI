import os, json, traceback
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, FileResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.conf import settings
from django.contrib import messages
from .models import QuestionPaper, Question, Student, CopySubmission, SubmissionPage, AnswerResult
from django.contrib.auth.models import User
from django.contrib.auth import login as auth_login


def home(request):
    if request.user.is_authenticated:
        from django.shortcuts import redirect
        return redirect('dashboard')
    from django.shortcuts import render
    return render(request, 'grader/home.html')
"""def register(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    if request.method == 'POST':"""
def register(request):
    """Teacher registration view"""
    if request.user.is_authenticated:
        return redirect('dashboard')

    if request.method == 'POST':
        first_name = request.POST.get('first_name', '').strip()
        last_name  = request.POST.get('last_name', '').strip()
        username   = request.POST.get('username', '').strip()
        email      = request.POST.get('email', '').strip()
        school     = request.POST.get('school', '').strip()
        password1  = request.POST.get('password1', '')
        password2  = request.POST.get('password2', '')

        # Validation
        if not all([first_name, last_name, username, email, password1]):
            messages.error(request, 'Sab fields fill karna zaroori hai.')
            return render(request, 'grader/register.html')

        if password1 != password2:
            messages.error(request, 'Passwords match nahi kar rahe. Dobara try karo.')
            return render(request, 'grader/register.html')

        if len(password1) < 8:
            messages.error(request, 'Password kam se kam 8 characters ka hona chahiye.')
            return render(request, 'grader/register.html')

        if User.objects.filter(username=username).exists():
            messages.error(request, f'Username "{username}" already liya ja chuka hai. Koi aur username choose karo.')
            return render(request, 'grader/register.html')

        if User.objects.filter(email=email).exists():
            messages.error(request, 'Yeh email already registered hai.')
            return render(request, 'grader/register.html')

        # Create user
        user = User.objects.create_user(
            username   = username,
            email      = email,
            password   = password1,
            first_name = first_name,
            last_name  = last_name,
        )
        # School name profile mein save karo (optional)
        # user.profile.school = school  # agar Profile model hai to
        user.save()

        # Auto login after registration
        auth_login(request, user)
        messages.success(request, f'Welcome {first_name}! Tumhara account ban gaya. Dashboard mein aagaye ho!')
        return redirect('dashboard')

    return render(request, 'grader/register.html')
@login_required
def dashboard(request):
    papers    = QuestionPaper.objects.filter(teacher=request.user).order_by('-created_at')
    subs      = CopySubmission.objects.filter(paper__teacher=request.user)
    recent    = subs.order_by('-submitted_at')[:5]
    done_subs = subs.filter(status='done')
    grade_counts = {'A+':0,'A':0,'B+':0,'B':0,'C':0,'F':0}
    for s in done_subs:
        g = s.grade()
        grade_counts[g] = grade_counts.get(g,0)+1
    stats = {
        'papers':       papers.count(),
        'students':     Student.objects.filter(teacher=request.user).count(),
        'total_subs':   subs.count(),
        'done_subs':    done_subs.count(),
        'avg_score':    round(sum(s.percentage for s in done_subs)/max(done_subs.count(),1),1),
        'pass_count':   sum(1 for s in done_subs if s.percentage>=33),
        'fail_count':   sum(1 for s in done_subs if s.percentage<33),
        'grade_counts': json.dumps(grade_counts),
    }
    return render(request,'grader/dashboard.html',{'papers':papers,'stats':stats,'recent':recent})


@login_required
def paper_create(request):
    if request.method=='POST':
        paper=QuestionPaper.objects.create(teacher=request.user,title=request.POST['title'],subject=request.POST['subject'],class_name=request.POST['class_name'])
        return redirect('paper_questions',pk=paper.pk)
    return render(request,'grader/paper_create.html')


@login_required
def paper_questions(request,pk):
    paper=get_object_or_404(QuestionPaper,pk=pk,teacher=request.user)
    if request.method=='POST':
        data=json.loads(request.body)
        paper.questions.all().delete()
        for i,q in enumerate(data['questions'],1):
            Question.objects.create(paper=paper,question_no=i,question_text=q['text'],q_type=q['type'],correct_answer=q['answer'],options=q.get('options',''),keywords=q.get('keywords',''),max_marks=float(q['marks']))
        paper.compute_total()
        return JsonResponse({'status':'ok','total':paper.total_marks})
    return render(request,'grader/paper_questions.html',{'paper':paper})


@login_required
def paper_list(request):
    papers=QuestionPaper.objects.filter(teacher=request.user).order_by('-created_at')
    return render(request,'grader/paper_list.html',{'papers':papers})


@login_required
def student_list(request):
    students=Student.objects.filter(teacher=request.user).order_by('name')
    return render(request,'grader/student_list.html',{'students':students})


@login_required
def student_create(request):
    if request.method=='POST':
        Student.objects.create(teacher=request.user,name=request.POST['name'],roll_number=request.POST['roll_number'],class_name=request.POST['class_name'],phone=request.POST.get('phone',''))
        messages.success(request,'Student added!')
        return redirect('student_list')
    return render(request,'grader/student_create.html')


#upload copy


@login_required
def upload_copy(request):
    steps=['Upload','Denoise','Skew fix','Threshold','Segment','OCR','Keywords','AI Check','Done']
    return render(request,'grader/upload_copy.html',{'papers':QuestionPaper.objects.filter(teacher=request.user),'students':Student.objects.filter(teacher=request.user),'steps':steps})
@login_required
@require_POST
def process_copy(request):
    paper_id=request.POST.get('paper_id')
    student_id=request.POST.get('student_id')
    images=request.FILES.getlist('pages')
    if not all([paper_id,student_id,images]):
        return JsonResponse({'error':'Paper, student aur images required'},status=400)
    paper=get_object_or_404(QuestionPaper,pk=paper_id,teacher=request.user)
    student=get_object_or_404(Student,pk=student_id,teacher=request.user)
    CopySubmission.objects.filter(student=student,paper=paper,status='failed').delete()
    submission=CopySubmission.objects.create(student=student,paper=paper,status='processing')
    try:
        page_paths=[]
        for i,img in enumerate(images,1):
            page=SubmissionPage.objects.create(submission=submission,page_number=i,image=img)
            page_paths.append(page.image.path)
        questions=list(paper.questions.all())
        print(f"[DEBUG] Pages:{len(page_paths)} Questions:{len(questions)}")
        if not questions:
            submission.status='done'; submission.total_marks_obtained=0; submission.percentage=0; submission.save()
            return JsonResponse({'status':'done','submission_id':submission.pk,'redirect':f'/result/{submission.pk}/'})
        from grader.dip.pipeline import run_dip_pipeline
        from grader.ocr.extractor import extract_all_answers
        from grader.ai_checker.evaluator import evaluate_full_submission
        all_extracted={}
        for path in page_paths:
            try:
                processed=run_dip_pipeline(path)
                extracted=extract_all_answers(processed,questions)
                print(f"[DEBUG] Extracted:{list(extracted.keys())}")
                all_extracted.update(extracted)
            except Exception as pe:
                print(f"[DEBUG] Page error:{pe}"); continue
        evaluate_full_submission(submission,all_extracted)
        submission.refresh_from_db()
        return JsonResponse({'status':'done','submission_id':submission.pk,'redirect':f'/result/{submission.pk}/'})
    except Exception as e:
        print(f"[ERROR]\n{traceback.format_exc()}")
        submission.status='failed'; submission.save()
        return JsonResponse({'error':str(e)[:200]},status=500)


@login_required
def result_view(request,pk):
    submission=get_object_or_404(CopySubmission,pk=pk,paper__teacher=request.user)
    results=submission.results.select_related('question').all()
    return render(request,'grader/result.html',{'submission':submission,'results':results,'grade':submission.grade()})


@login_required
def download_pdf(request,pk):
    submission=get_object_or_404(CopySubmission,pk=pk,paper__teacher=request.user)
    from grader.utils.helpers import generate_result_pdf
    pdf_dir=os.path.join(settings.MEDIA_ROOT,'results'); os.makedirs(pdf_dir,exist_ok=True)
    pdf_path=os.path.join(pdf_dir,f'result_{pk}.pdf')
    generate_result_pdf(submission,pdf_path)
    return FileResponse(open(pdf_path,'rb'),as_attachment=True,filename=f'result_{submission.student.name}_{submission.paper.subject}.pdf')


@login_required
def send_whatsapp(request,pk):
    submission=get_object_or_404(CopySubmission,pk=pk,paper__teacher=request.user)
    from grader.utils.helpers import send_whatsapp_result
    success,msg=send_whatsapp_result(submission)
    return JsonResponse({'success':success,'message':msg})


@login_required
def all_results(request):
    paper_id=request.GET.get('paper')
    subs=CopySubmission.objects.filter(paper__teacher=request.user,status='done').select_related('student','paper').order_by('-submitted_at')
    if paper_id: subs=subs.filter(paper_id=paper_id)
    papers=QuestionPaper.objects.filter(teacher=request.user)
    scores=[s.percentage for s in subs]
    return render(request,'grader/all_results.html',{'submissions':subs,'papers':papers,'selected_paper':paper_id,'avg_score':round(sum(scores)/max(len(scores),1),1),'topper':subs.order_by('-percentage').first(),'pass_count':sum(1 for p in scores if p>=33),'fail_count':sum(1 for p in scores if p<33)})


#FEATURE 1: Excel Export
@login_required
def export_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    paper_id=request.GET.get('paper')
    subs=CopySubmission.objects.filter(paper__teacher=request.user,status='done').select_related('student','paper').order_by('student__roll_number')
    if paper_id: subs=subs.filter(paper_id=paper_id)
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Results"
    hfill=PatternFill("solid",fgColor="1A56DB"); hfont=Font(bold=True,color="FFFFFF",size=11); ctr=Alignment(horizontal='center',vertical='center')
    headers=["Roll No","Student Name","Subject","Class","Total Marks","Obtained","Percentage","Grade","Result"]
    for col,h in enumerate(headers,1):
        cell=ws.cell(row=1,column=col,value=h); cell.font=hfont; cell.fill=hfill; cell.alignment=ctr
    for row,s in enumerate(subs,2):
        data=[s.student.roll_number,s.student.name,s.paper.subject,s.student.class_name,s.paper.total_marks,s.total_marks_obtained,f"{s.percentage}%",s.grade(),"Pass" if s.percentage>=33 else "Fail"]
        for col,val in enumerate(data,1):
            cell=ws.cell(row=row,column=col,value=val); cell.alignment=ctr
            if col==9: cell.font=Font(color="057A55" if val=="Pass" else "C81E1E",bold=True)
    for col in range(1,len(headers)+1):
        ws.column_dimensions[get_column_letter(col)].width=18
    response=HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition']='attachment; filename="AutoGrade_Results.xlsx"'
    wb.save(response); return response


#FEATURE 2: Class Analytics
@login_required
def class_analytics(request):
    paper_id=request.GET.get('paper')
    papers=QuestionPaper.objects.filter(teacher=request.user)
    subs=CopySubmission.objects.filter(paper__teacher=request.user,status='done').select_related('student','paper')
    if paper_id: subs=subs.filter(paper_id=paper_id)
    grade_dist={'A+':0,'A':0,'B+':0,'B':0,'C':0,'F':0}
    scores=[]
    for s in subs:
        g=s.grade(); grade_dist[g]=grade_dist.get(g,0)+1; scores.append(s.percentage)
    avg=round(sum(scores)/max(len(scores),1),1)
    q_perf={}
    for s in subs:
        for r in s.results.select_related('question').all():
            qkey=f"Q{r.question.question_no}: {r.question.question_text[:35]}"
            if qkey not in q_perf: q_perf[qkey]={'total':0,'count':0}
            q_perf[qkey]['total']+=r.final_marks; q_perf[qkey]['count']+=1
    q_avg={k:round(v['total']/max(v['count'],1),1) for k,v in q_perf.items()}
    return render(request,'grader/analytics.html',{'papers':papers,'sel_paper':paper_id,'grade_dist':json.dumps(grade_dist),'avg':avg,'high':round(max(scores,default=0),1),'low':round(min(scores,default=0),1),'passed':sum(1 for p in scores if p>=33),'failed':len(scores)-sum(1 for p in scores if p>=33),'total':len(scores),'q_avg':json.dumps(q_avg),'topper':subs.order_by('-percentage').first()})
#FEATURE 3: Student Portal
def student_result_portal(request):
    result=None; error=None; submitted=False
    if request.method=='POST':
        roll=request.POST.get('roll_number','').strip(); phone=request.POST.get('phone','').strip(); submitted=True
        try:
            student=Student.objects.get(roll_number=roll,phone=phone)
            subs=CopySubmission.objects.filter(student=student,status='done').order_by('-submitted_at')
            result=subs if subs.exists() else None
            if not result: error="Koi result nahi mila. Teacher se milein."
        except Student.DoesNotExist:
            error="Roll number ya phone number galat hai."
    return render(request,'grader/student_portal.html',{'result':result,'error':error,'submitted':submitted})


#FEATURE 4: Bulk PDF ZIP Download
@login_required
def bulk_pdf_download(request):
    import zipfile, io
    from grader.utils.helpers import generate_result_pdf
    paper_id=request.GET.get('paper')
    if not paper_id: return HttpResponse("Paper ID required",status=400)
    subs=CopySubmission.objects.filter(paper_id=paper_id,paper__teacher=request.user,status='done').select_related('student','paper')
    zip_buf=io.BytesIO(); pdf_dir=os.path.join(settings.MEDIA_ROOT,'results'); os.makedirs(pdf_dir,exist_ok=True)
    with zipfile.ZipFile(zip_buf,'w',zipfile.ZIP_DEFLATED) as zf:
        for s in subs:
            pp=os.path.join(pdf_dir,f'result_{s.pk}.pdf')
            generate_result_pdf(s,pp); zf.write(pp,f"{s.student.roll_number}_{s.student.name}.pdf")
    zip_buf.seek(0)
    response=HttpResponse(zip_buf,content_type='application/zip')
    response['Content-Disposition']='attachment; filename="All_Results.zip"'
    return response